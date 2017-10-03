##---------------------------------------------------------------------
## Program: geocoder.py
## Version: 1.0
##
## Author: Nick Joslyn
## Institution: Simpson College
##
## Purpose: 
##	Use GoogleMaps' geocoding ability to extract latitude and longitude
##	for a given set of addresses.
##
## Instructions:
##	Black box Python script. Generalized and runs on user-input.
##	If you desire, you can edit the code. It is thoroughly commented.
##	
##	You must have openpyxl and the Google Maps modules installed.
##	Additionally, you must have a Google Maps API key. API keys have
##	quotas. 2500 requests per day. Only 100 at a time. If you enable
##	billing, your quotas increase, and you should increase the apiSingleQuota
##	variable accordingly.
##
## Notes: 
##	If Google cannot find the address, the code is written to put a
##	blank cell in the output file. If you find a blank cell in your
##	output file, then you must manually decide how to deal with that
##	particular data point. This will not happen often, and the program
##	will tell you which row was left blank.
##
##---------------------------------------------------------------------

##=====================================================================

## This section uses user input to determine the six important variables
## that instruct the rest of the program on how to run.

# Input is an Excel file with addresses to be geocoded
inputFileName = input("Type the name of the Excel file containing addresses (Specify .xlsx extension.): ") 

# This allows the program to skip over any header information rows
headerOffset = int(input("What row does your data start on? (Specify numerical value. Likely 2.): "))

# This indicates to the program which column in the Excel file contains the addresses
columnAddress = int(input("What column number are the addresses in? (Specify numerical value.): "))

# This is the API Key of the user to allow Google Maps information extraction
apiKey = input("What is your Google Maps API Key?: ")

# This is the Excel file where the geocoded results will be written to
outputFileName = input("Type the name of an Excel file you want the results printed to (Specify .xlsx extension. New file recommended): ")

# This is not a user input driven variable.
# However, it is included in this section because it could be changed.
# If you enabled billing, this could be much higher.
# Without enabling billing, 100 is the maximum amount allowed by Google in one call.
#apiSingleQuota = 100

##=====================================================================

## From here on, the program runs in a self-contained manner.
## Everything is based on the user input variables.

print("\n----------------Program Running----------------\n")

# Module for reading/writing from Excel Files
# Module for timing how long the program takes
# Modules for extracting Google Maps information

import openpyxl
from openpyxl import Workbook

import time

import googlemaps
from datetime import datetime

# Begin Timer
start = time.time()

# Load the file we want to read from and set sheet to hold the information.
book = openpyxl.load_workbook(inputFileName)
sheet = book.active

# Count the number of addresses to be geocoded.
# We subtract the headerOffset (minus 1) to subtract out rows 
# that are not demand nodes.
numberOfAddresses = sheet.max_row - (headerOffset - 1)

# Loop through the spreadsheet, filling up addressList with
# the address of each in sequential order.
# Note, we add the headerOffset variable to the loop so we start with the correct row
addressList = []
for i in range(numberOfAddresses):
		addressList.append(sheet.cell(row = (i+headerOffset), column = columnAddress).value)

# Set the active worksheet to begin the write-out process
wb = Workbook()
ws = wb.active

# Supply header information for the two columns that we will write
# to in the output file
ws.cell(row = 1, column = 1, value = "Address")
ws.cell(row = 1, column = 2, value = "Latitude")
ws.cell(row = 1, column = 3, value = "Longitude")

gmaps = googlemaps.Client(key = apiKey)

for i in range(numberOfAddresses):
	geocode_result = gmaps.geocode(addressList[i])
	try:
		ws.cell(row = (i + 2), column = 1, value = addressList[i])
		ws.cell(row = (i + 2), column = 2, value = geocode_result[0]['geometry']['location']['lat'])
		ws.cell(row = (i + 2), column = 3, value = geocode_result[0]['geometry']['location']['lng'])
	except:
		print("Error. The value could not be found. Row: " + str(i + 2))
	if (i%1000 == 0):
		print("Completed " + str(i) + " addresses.")
    
# Save the output Excel file with the user-inputted name
wb.save(outputFileName)

print("\n----------------Program Done----------------\n")

print("Results written to " + str(outputFileName) + "\n")

# End Timer
# Print elapsed time rounded to two places
end = time.time()
print("Time for program to complete: " + str(round(end - start, 2)))
